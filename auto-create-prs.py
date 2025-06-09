import os
import time
import math
import csv
import logging
import subprocess
import openpyxl
import requests
from datetime import datetime
from zoneinfo import ZoneInfo
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from requests.exceptions import RequestException
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import webbrowser

# ------------------- Configuration -------------------
GITHUB_TOKEN = "your_token_here"  # Replace with your GitHub Token
INPUT_PATH = r"C:\Users\v-bowenyang\Desktop\Daily_Publishing\OPS-Publish-10_00.xlsx"
PRE_LINKS_FILE = r"C:\Users\v-bowenyang\Desktop\Daily_Publishing\Sync_PR\OPS-Publish-10_00.csv"
BASE_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "PR_created_result")

# Directory to save output results (log + Excel)
BASE_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "PR_created_result")

# --------------------- Setup Output Paths ---------------------
current_date = datetime.now().strftime("%Y%m%d")
safe_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
daily_dir = os.path.join(BASE_OUTPUT_DIR, current_date)
OUTPUT_DIR = os.path.join(daily_dir, f"run_{safe_timestamp}")
os.makedirs(OUTPUT_DIR, exist_ok=True)
log_path = os.path.join(OUTPUT_DIR, f"log_{safe_timestamp}.txt")

# --------------------- Logging Configuration ------------------
logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format="[%(asctime)s] %(message)s",
    datefmt="%H:%M:%S",
)

# --------------------- GitHub Session -------------------------
def get_session():
    """
    Create a pre-configured requests.Session for communicating with GitHub API.

    Features:
    - Automatically retries failed requests (e.g., temporary 5xx errors) up to 2 times.
    - Attaches GitHub personal access token for authenticated access.

    Returns:
        requests.Session: Authenticated session ready for GitHub API requests.
    """
    session = requests.Session()

    # Configure retry strategy: retry failed HTTPS requests up to 2 times
    retry_adapter = HTTPAdapter(max_retries=2)
    session.mount("https://", retry_adapter)

    # Attach GitHub token to the session header for authentication
    session.headers.update({
        "Authorization": f"token {GITHUB_TOKEN}"
    })

    return session

# --------------------- PR Title Generator ---------------------
def get_pr_title():
    """
    Generate a pull request title based on current US Pacific time.

    Format:
        {month}/{day}/{year} {AM/PM} Publish

    Example:
        6/9/2025 AM Publish

    Returns:
        str: A timestamped PR title indicating the publish batch.
    """
    # Get the current time in US Pacific timezone (auto handles PST/PDT)
    now_pacific = datetime.now(ZoneInfo("America/Los_Angeles"))

    # Determine AM/PM marker based on hour
    hour = now_pacific.hour
    meridiem = "AM" if hour < 12 else "PM"

    # Construct the PR title string
    return f"{now_pacific.month}/{now_pacific.day}/{now_pacific.year} {meridiem} Publish"

# --------------------- Parse Compare Links --------------------
def parse_compare_link(link):
    """
    Parse a GitHub compare URL and extract organization, repo, base, and head branch names.

    Expected format:
        https://github.com/{org}/{repo}/compare/{base}...{head}

    Args:
        link (str): GitHub compare URL.

    Returns:
        tuple: (org, repo, base, head)

    Raises:
        ValueError: If the link is malformed or missing expected components.
    """
    try:
        # Use urlparse to extract the path portion of the URL
        parsed = urlparse(link)
        parts = parsed.path.strip("/").split("/")

        # Check basic structure validity
        if len(parts) < 4 or parts[2] != "compare":
            raise ValueError(f"Invalid compare link format: {link}")

        # Extract org, repo, and compare segment
        org, repo, _, compare_part = parts[:4]

        # Split base and head branches using '...'
        base, head = compare_part.split("...")[0:2]

        return org, repo, base, head

    except Exception:
        # Raise explicit error if parsing fails
        raise ValueError(f"Invalid compare link format: {link}")

# --------------------- Check Existing PRs ---------------------
# Check if a pull request already exists between the specified base and head branches
def check_existing_pr(session, org, repo, base, head):
    """
    Query GitHub to check if a pull request already exists between base and head branches.

    Args:
        session (requests.Session): Authenticated GitHub session.
        org (str): GitHub organization or username.
        repo (str): Repository name.
        base (str): Base branch name (e.g., main).
        head (str): Head branch name (e.g., feature/new-ui).

    Returns:
        str or None: URL of the existing pull request if found; otherwise None.
    """
    url = f"https://api.github.com/repos/{org}/{repo}/pulls"
    params = {"state": "open"}

    # Fetch all open PRs in the repository
    response = session.get(url, params=params)
    if response.status_code != 200:
        logging.warning(f"GitHub PR check failed with status {response.status_code} for {org}/{repo}")
        return None

    pulls = response.json()

    # Check each open PR to see if it matches the base and head
    for pr in pulls:
        if pr.get("base", {}).get("ref") == base and pr.get("head", {}).get("ref") == head:
            return pr.get("html_url")  # Matching PR found

    return None  # No matching PR found

# --------------------- Create Pull Request ---------------------
# Attempt to create a pull request on GitHub, with retry and error handling
def create_pull_request(session, org, repo, base, head, title, max_retries=3):
    """
    Create a new pull request via GitHub API, with retry logic and detailed error handling.

    Args:
        session (requests.Session): Authenticated GitHub session.
        org (str): GitHub organization name.
        repo (str): Repository name.
        base (str): Base branch name (e.g., main).
        head (str): Head branch name (e.g., feature).
        title (str): Pull request title.
        max_retries (int): Max retry attempts for transient server errors.

    Returns:
        tuple: (status, pr_link, reason)
            - status: "Created", "Duplicate", or "Error"
            - pr_link: URL to the created PR (or "-" if not available)
            - reason: Reason string for failure or duplication
    """
    url = f"https://api.github.com/repos/{org}/{repo}/pulls"
    data = {"title": title, "head": head, "base": base}

    for attempt in range(max_retries):
        response = session.post(url, json=data)

        # ---- PR successfully created
        if response.status_code == 201:
            return "Created", response.json()["html_url"], ""

        # ---- PR already exists (422 Unprocessable Entity)
        elif response.status_code == 422:
            pr_url = check_existing_pr(session, org, repo, base, head)
            return (
                "Duplicate",
                pr_url if pr_url else "-",
                "Pull request already exists." if pr_url else "PR exists but URL not found.",
            )

        # ---- Temporary server error, retry
        elif response.status_code in [500, 502, 503, 504]:
            logging.warning(f"Server error ({response.status_code}) on attempt {attempt + 1}. Retrying...")
            time.sleep(2)
            continue

        # ---- Not found (likely invalid repo or branch)
        elif response.status_code == 404:
            return "Error", "-", "Repository or branch not found."

        # ---- Forbidden (rate limited or permission issue)
        elif response.status_code == 403:
            return "Error", "-", "Rate limit exceeded or token permissions insufficient."

        # ---- Unauthorized (token expired or invalid)
        elif response.status_code == 401:
            return "Error", "-", "Invalid GitHub token."

        # ---- Other unexpected error
        else:
            return "Error", "-", f"Unexpected error: {response.status_code}"

    # ---- All retry attempts failed
    return "Error", "-", "Failed after max retries due to server error."

# --------------------- Load Links from Excel --------------------
def load_compare_links(path):
    """Read compare links from the first column of an Excel file."""
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    links = [row[0].strip() for row in sheet.iter_rows(min_row=2, values_only=True)
             if row[0] and isinstance(row[0], str) and row[0].startswith("http")]
    return links

# --------------------- Process Individual Link ------------------
def process_link(link, pr_title):
    """
    Process a single GitHub compare link to determine if a pull request should be created.

    Steps:
    1. Parse link to extract repo and branch info.
    2. Check commit diff from GitHub API.
    3. If there are commits, attempt to create a PR.
    4. Return detailed result for logging and reporting.

    Args:
        link (str): GitHub compare link (e.g., https://github.com/org/repo/compare/base...head)
        pr_title (str): Title to use when creating the pull request.

    Returns:
        list: [link, status, pr_link, commits, files_changed, reason]
    """
    session = get_session()
    logging.info(f"Processing: {link}")

    try:
        # ---- Parse the compare link to extract org/repo/base/head
        org, repo, base, head = parse_compare_link(link)

        # ---- Construct API URL to fetch comparison info
        compare_url = f"https://api.github.com/repos/{org}/{repo}/compare/{base}...{head}"
        compare_resp = session.get(compare_url)

        # ---- If compare endpoint is invalid (e.g. deleted branch), return error
        if compare_resp.status_code == 404:
            return [link, "Error", "-", "-", "-", "Compare link not valid."]

        # ---- Parse commit and file change data
        data = compare_resp.json()
        commits = data.get("total_commits", "-")
        files_changed_raw = data.get("files", [])
        files_changed = len(files_changed_raw)

        # GitHub truncates file list after 300 changes — indicate overflow
        if files_changed == 300:
            files_changed = "300+"

        # ---- Skip PR creation if there are no new commits
        if commits == 0:
            return [link, "Skipped", "-", commits, files_changed, "No new commits to publish."]

        # ---- Attempt to create the pull request
        status, pr_link, reason = create_pull_request(session, org, repo, base, head, pr_title)
        return [link, status, pr_link or "-", commits, files_changed, reason]

    except Exception as e:
        # ---- On any unexpected error, log and return as failed entry
        logging.exception(f"Error processing {link}:")
        return [link, "Error", "-", "-", "-", str(e)]

# ---------------- Save PR Results to Excel ------------------
def save_results_to_excel(results, output_path):
    """
    Save the PR creation results to an Excel file with headers, auto column sizing, and font formatting.

    Args:
        results (List[List[str]]): A list of rows containing PR creation results.
        output_path (str): Path to the output Excel file.
    """
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PR Results"

    # Define and write header row
    headers = ["Compare Link", "Result", "PR Link", "Commits", "Files Changed", "Reason"]
    ws.append(headers)

    # Set font styles
    header_font = Font(name="Segoe UI", size=11, bold=True)
    cell_font = Font(name="Segoe UI", size=11)

    # Apply header font style
    for cell in ws[1]:
        cell.font = header_font

    # Write data rows
    for row in results:
        ws.append(row)

    # Auto-adjust column widths and apply font style
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2
        for cell in col:
            cell.font = cell_font

    # Save workbook with fallback handling if the file is open
    try:
        wb.save(output_path)
    except PermissionError:
        fallback = os.path.splitext(output_path)[0] + "_retry.xlsx"
        print()
        print(f"❌ Unable to save to {output_path}. File might be open. Saving as: {fallback}")
        wb.save(fallback)
        output_path = fallback

    # Automatically open the file in the default Excel viewer
    webbrowser.open(output_path)
    logging.info(f"Results saved to: {output_path}")

# -------- Open Links in Edge Browser (Grouped Tabs) --------
def open_links_in_edge_window_grouped(links, group_size=15):
    """
    Open a list of URLs in Microsoft Edge, dividing them into multiple windows
    with a specified number of tabs per window to avoid overload.

    Args:
        links (list): List of URLs to open.
        group_size (int): Approximate number of tabs per browser window.
    """
    # Path to Microsoft Edge executable (adjust if installed elsewhere)
    edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"

    # Validate Edge installation path
    if not os.path.exists(edge_path):
        print()
        print(f"❌ Microsoft Edge not found at: {edge_path}")
        return

    total = len(links)
    if total == 0:
        return  # No links to open

    # Calculate number of browser windows and actual group size
    num_groups = math.ceil(total / group_size)
    group_size = math.ceil(total / num_groups)  # Redistribute evenly

    print()
    print(f"🌐 Opening {total} links in {num_groups} Edge windows (~{group_size} tabs each)...")

    # Launch Edge windows in batches
    for i in range(0, total, group_size):
        group = links[i:i + group_size]
        subprocess.Popen([edge_path, "--new-window"] + group)
        time.sleep(1)  # Small delay to prevent browser overload

# -------------- Load Pre-check Links ------------------------
def open_links_from_excel(path):
    """
    Load pre-check links from a CSV or XLSX file and prompt the user to open them in Edge browser.

    Args:
        path (str): Path to the pre-check file (CSV or XLSX format).
    """
    links = []
    try:
        # ---- Read from CSV file ----
        if path.lower().endswith(".csv"):
            with open(path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header row if present
                links = [row[0].strip() for row in reader if row and row[0].startswith("http")]

        # ---- Read from Excel file (.xlsx) ----
        else:
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
            links = [row[0].strip() for row in sheet.iter_rows(min_row=2, values_only=True)
                     if row[0] and row[0].startswith("http")]

        # ---- Handle empty or invalid file ----
        if not links:
            print("ℹ️ No valid links found in pre-check file.")
            return

        # ---- Ask for user confirmation before opening ----
        confirm = input(f"\n📘 {len(links)} pre-check links found. Open in Edge? (y/n): ").strip().lower()
        if confirm == "y":
            open_links_in_edge_window_grouped(links, group_size=15)
        else:
            print()
            print("❌ Skipped opening pre-check links.")

    except Exception as e:
        # ---- Error handling during file reading ----
        print()
        print(f"❌ Failed to open links from file: {e}")

# -------------- Open PR Links in Browser --------------------
def open_pr_links_in_browser(results):
    """
    Extract successfully created or duplicate PR links from results and open them in Edge.

    Args:
        results (list): List of PR processing results with columns [link, status, PR URL, ...]
    """
    # Filter out links where the PR was successfully created or already existed
    pr_links = [
        row[2]
        for row in results
        if row[1] in ("Created", "Duplicate") and row[2].startswith("http")
    ]

    # If no links to open, notify user and return
    if not pr_links:
        print("\nℹ️ No PR links to open.")
        return

    # Ask for confirmation before opening links in browser
    confirm = input(f"\n🔗 {len(pr_links)} PR links found. Open in Edge? (y/n): ").strip().lower()
    if confirm == "y":
        open_links_in_edge_window_grouped(pr_links, group_size=7)
    else:
        print()
        print("❌ Skipped opening PR links.")

# -------------- Save Results and Summarize --------------------
def summarize_and_save_results(results, stats, output_dir, timestamp):
    """
    Summarize the pull request creation results, export them to an Excel file,
    and optionally open successful PR links in the browser.

    Args:
        results (list): List of individual PR processing results.
        stats (dict): Aggregated status counts (e.g., Created, Skipped, Error).
        output_dir (str): Directory where the Excel file will be saved.
        timestamp (str): Timestamp string used to name the output file.
    """
    # Print and log summary statistics
    print(f"\n📊 Summary: {stats['Created']} Created, {stats['Skipped']} Skipped, "
          f"{stats['Duplicate']} Duplicate, {stats['Error']} Error")
    logging.info(f"Summary: {stats}")

    # Compose Excel file path using timestamp and save all results
    output_file = os.path.join(output_dir, f"pr_creation_results_{timestamp}.xlsx")
    save_results_to_excel(results, output_file)

    # Ask user whether to open all successfully created PR links in browser
    open_pr_links_in_browser(results)

# -------------- Process All Compare Links ---------------------
def process_all_links(compare_links, pr_title):
    """
    Process all GitHub compare links in parallel using ThreadPoolExecutor,
    and collect PR creation results with status statistics.
    
    Args:
        compare_links (list[str]): List of GitHub compare URLs to process.
        pr_title (str): Title to use when creating pull requests.
    
    Returns:
        results (list[list]): Each row contains [link, status, PR link, commits, files changed, reason].
        stats (dict): Count of outcomes: Created, Skipped, Duplicate, Error.
    """
    results = []
    stats = {"Created": 0, "Skipped": 0, "Duplicate": 0, "Error": 0}

    # Initialize a thread pool with up to 5 concurrent workers
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Submit each compare link as a separate task
        futures = [executor.submit(process_link, link, pr_title) for link in compare_links]

        # Collect results as they complete
        for idx, future in enumerate(as_completed(futures), 1):
            result = future.result()
            results.append(result)

            # Increment the appropriate status counter
            stats[result[1]] += 1

            # Print progress every 5 links or at the end
            if idx % 5 == 0 or idx == len(compare_links):
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Processed {idx}/{len(compare_links)} links...")

    return results, stats

# -------------- Confirm and Prepare Output --------------------
# Prepare output directory, initialize logging, and confirm whether to proceed
def confirm_run_and_prepare_output():
    """
    Prompt user for confirmation before proceeding with PR creation.
    If confirmed, prepare a timestamped output directory and configure logging.

    Returns:
        confirmed (bool): True if user confirms to proceed, False otherwise.
        output_dir (str|None): Directory path to save results and logs.
        timestamp (str|None): Timestamp string used for naming output files.
    """
    # Show user the file to be processed and ask for confirmation
    print(f"\n📄 File to process: {INPUT_PATH}")
    if input("⚠️ Confirm to start PR creation for this file? (y/n): ").strip().lower() != "y":
        print("❌ Cancelled by user.")
        return False, None, None

    # Generate a unique timestamp and output path for this run
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")  # e.g. 20250609
    timestamp = now.strftime("%Y%m%d_%H%M%S")  # e.g. 20250609_103045
    output_dir = os.path.join(BASE_OUTPUT_DIR, date_str, f"run_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    # Configure logging to a timestamped log file inside the output directory
    log_path = os.path.join(output_dir, f"log_{timestamp}.txt")
    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format="[%(asctime)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    return True, output_dir, timestamp

# ------------------------ Main Logic ------------------------
def main():
    """Main routine for pre-check, PR creation, and result reporting."""
    
    # Step 1: Open pre-check links (e.g., sync PRs) before proceeding
    open_links_from_excel(PRE_LINKS_FILE)

    # Step 2: Confirm with user and prepare output/log paths
    confirmed, output_dir, timestamp = confirm_run_and_prepare_output()
    if not confirmed:
        return  # Exit if user cancels

    # Step 3: Generate standardized PR title with timestamp (PST)
    pr_title = get_pr_title()

    # Step 4: Load compare links from input Excel file
    compare_links = load_compare_links(INPUT_PATH)

    # Step 5: Notify start of PR creation process
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 🚀 Starting PR creation with title: {pr_title}")

    # Step 6: Process all compare links concurrently and track results
    results, stats = process_all_links(compare_links, pr_title)

    # Step 7: Print summary, save Excel output, and open PR links in browser
    summarize_and_save_results(results, stats, output_dir, timestamp)

# ------------------------ Entry Point ------------------------
if __name__ == "__main__":
    main()
